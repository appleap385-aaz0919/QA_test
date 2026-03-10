# -*- coding: utf-8 -*-
"""
테스트 결과를 엑셀에 저장
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json

# 엑셀 파일 경로
EXCEL_PATH = "D:/jaehyuk.myung/claude_demo/Demo_13_QA/Quick_TC_list/AIDT 모니터링_운영계_점검 리스트.xlsx"

# 테스트 결과
test_results = {
    "teacher": {
        "홈_기본노출": "FAIL",
        "홈_출석체크": "FAIL",
        "홈_과제": "FAIL",
        "교과서_수업시작": "FAIL",
        "과제_목록": "FAIL",
        "AI평가_기본": "FAIL",
        "AI학습관_기본": "PASS",
        "오답노트_기본": "PASS",
        "학생_현황": "PASS",
        "내자료_기본": "PASS",
        "AI보조교사_버튼": "PASS"
    },
    "student": {
        "홈_기본노출": "FAIL",
        "홈_기분설정": "PASS",
        "교과서_기본": "PASS",
        "과제_목록": "PASS",
        "모둠활동_기본": "PASS",
        "AI평가_기본": "PASS",
        "AI학습관_기본": "PASS",
        "오답노트_기본": "PASS",
        "학습결과_기본": "PASS"
    }
}

# 오늘 날짜
today = datetime.now().strftime("%Y-%m-%d")

# 엑셀 파일 로드
wb = load_workbook(EXCEL_PATH)

# 새 시트 생성 (이름: 오늘 날짜)
sheet_name = today
if sheet_name in wb.sheetnames:
    # 이미 존재하면 삭제 후 재생성
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name)

# 스타일 정의
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 타이틀
ws['A1'] = f"AIDT 모니터링 QA 테스트 결과 ({today})"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')

# 테스트 정보
ws['A3'] = "Test 목적"
ws['B3'] = "AIDT LMS 서비스 기본 동작 점검"
ws['A4'] = "Tester"
ws['B4'] = "자동화 테스트"
ws['A5'] = "서버"
ws['B5'] = "운영계"
ws['A6'] = "URL"
ws['B6'] = "https://www.aidt.ai/lms-web/dev/entry-aidt-2025?school=m&subject=eng&grade=2&semester=all&authorName=yoon"

# 헤더 행
row = 8
headers = ['구분', '과목', '권한', 'Step1', 'Step2', '경로', '검증 결과', '오류 현상 및 비고']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 교사 테스트 결과
row += 1
ws.cell(row=row, column=1, value="교사").font = Font(bold=True)
ws.cell(row=row, column=3, value="교사").font = Font(bold=True)

teacher_items = [
    ("홈", None, "수업 정보, 교과서, 출석 체크, 과제 노출 확인", test_results["teacher"].get("홈_기본노출", "NT")),
    ("홈", None, "출석 체크 기능 확인", test_results["teacher"].get("홈_출석체크", "NT")),
    ("홈", None, "과제 목록 노출 확인", test_results["teacher"].get("홈_과제", "NT")),
    ("교과서", None, "수업 시작하기 버튼 > 수업 뷰어 > 콘텐츠 노출 확인", test_results["teacher"].get("교과서_수업시작", "NT")),
    ("과제", "과제", "과제 리스트 노출 확인", test_results["teacher"].get("과제_목록", "NT")),
    ("AI 평가", None, "평가 목록 및 배포 기능 확인", test_results["teacher"].get("AI평가_기본", "NT")),
    ("AI 학습관", None, "AI 학습관 진입 확인", test_results["teacher"].get("AI학습관_기본", "NT")),
    ("오답 노트", None, "오답 노트 진입 확인", test_results["teacher"].get("오답노트_기본", "NT")),
    ("학생", "학급 분석", "학생 현황 데이터 확인", test_results["teacher"].get("학생_현황", "NT")),
    ("내 자료", None, "내 자료 진입 확인", test_results["teacher"].get("내자료_기본", "NT")),
    ("AI 보조교사", None, "챗봇 버튼 노출 확인", test_results["teacher"].get("AI보조교사_버튼", "NT")),
]

for step1, step2, path, result in teacher_items:
    row += 1
    ws.cell(row=row, column=4, value=step1).border = thin_border
    ws.cell(row=row, column=5, value=step2).border = thin_border
    ws.cell(row=row, column=6, value=path).border = thin_border
    result_cell = ws.cell(row=row, column=7, value=result)
    result_cell.border = thin_border
    result_cell.alignment = Alignment(horizontal='center')
    if result == "PASS":
        result_cell.fill = pass_fill
    elif result == "FAIL":
        result_cell.fill = fail_fill
        ws.cell(row=row, column=8, value="자동화 테스트: 요소 찾기 실패").border = thin_border

# 학생 테스트 결과
row += 2
ws.cell(row=row, column=1, value="학생").font = Font(bold=True)
ws.cell(row=row, column=3, value="학생").font = Font(bold=True)

student_items = [
    ("홈", None, "홈 화면 기본 노출 확인", test_results["student"].get("홈_기본노출", "NT")),
    ("홈", None, "기분 설정 확인", test_results["student"].get("홈_기분설정", "NT")),
    ("교과서", None, "교과서 보기 진입 확인", test_results["student"].get("교과서_기본", "NT")),
    ("과제", "과제", "과제 리스트 노출 확인", test_results["student"].get("과제_목록", "NT")),
    ("모둠 활동", None, "모둠 활동 진입 확인", test_results["student"].get("모둠활동_기본", "NT")),
    ("AI 평가", None, "AI 평가 진입 확인", test_results["student"].get("AI평가_기본", "NT")),
    ("AI 학습관", None, "AI 학습관 진입 확인", test_results["student"].get("AI학습관_기본", "NT")),
    ("오답 노트", None, "오답 노트 진입 확인", test_results["student"].get("오답노트_기본", "NT")),
    ("학습 결과", None, "학습 결과 진입 확인", test_results["student"].get("학습결과_기본", "NT")),
]

for step1, step2, path, result in student_items:
    row += 1
    ws.cell(row=row, column=4, value=step1).border = thin_border
    ws.cell(row=row, column=5, value=step2).border = thin_border
    ws.cell(row=row, column=6, value=path).border = thin_border
    result_cell = ws.cell(row=row, column=7, value=result)
    result_cell.border = thin_border
    result_cell.alignment = Alignment(horizontal='center')
    if result == "PASS":
        result_cell.fill = pass_fill
    elif result == "FAIL":
        result_cell.fill = fail_fill
        ws.cell(row=row, column=8, value="자동화 테스트: 요소 찾기 실패").border = thin_border

# 컬럼 너비 조정
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 30

# 저장
wb.save(EXCEL_PATH)
print(f"엑셀 저장 완료: {EXCEL_PATH}")
print(f"새 시트 이름: {sheet_name}")
