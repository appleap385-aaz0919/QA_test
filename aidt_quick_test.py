# -*- coding: utf-8 -*-
"""
AIDT LMS QA 테스트 - window.open 패치 적용
"""

import asyncio
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ENTRY_URL = "https://www.aidt.ai/lms-web/dev/entry-aidt-2025?school=m&subject=eng&grade=2&semester=all&authorName=yoon"
EXCEL_PATH = "D:/jaehyuk.myung/claude_demo/Demo_13_QA/Quick_TC_list/AIDT 모니터링_운영계_점검 리스트.xlsx"
SCREENSHOT_DIR = "D:/jaehyuk.myung/claude_demo/Demo_13_QA/screenshots"

results = {"teacher": {}, "student": {}}

async def screenshot(page, name):
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    await page.screenshot(path=f"{SCREENSHOT_DIR}/{name}.png", full_page=True)

async def check(page, text, timeout=3000):
    try:
        await page.wait_for_selector(f"text={text}", timeout=timeout, state="visible")
        return True
    except:
        return False

async def click_tab(page, tab):
    try:
        await page.click(f"text={tab}", timeout=5000)
        await page.wait_for_timeout(1500)
        return True
    except:
        return False

async def patch_window_open(page):
    """window.open을 window.location.href로 패치"""
    await page.evaluate("""
        window._originalOpen = window.open;
        window.open = function(url, target, features) {
            if (url && !url.startsWith('javascript')) {
                window.location.href = url;
            }
            return null;
        };
    """)

async def run_tests():
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=500)
        context = await browser.new_context(viewport={"width": 1920, "height": 1080})
        page = await context.new_page()

        # Entry
        print("1. Entry 페이지 접속...")
        await page.goto(ENTRY_URL, timeout=60000)
        await page.wait_for_timeout(3000)

        # window.open 패치
        await patch_window_open(page)

        await screenshot(page, "v2_entry")

        # 선생님 선택
        print("2. 선생님 선택...")
        try:
            # 다양한 선택자 시도
            selectors = [
                "text=선생님",
                "text=교사",
                "text=상담사",
                "button:has-text('선생님')",
                "[class*='teacher']",
            ]
            for sel in selectors:
                try:
                    if await page.is_visible(sel, timeout=2000):
                        await page.click(sel)
                        print(f"  클릭 성공: {sel}")
                        break
                except:
                    continue

            await page.wait_for_timeout(5000)
        except Exception as e:
            print(f"  선생님 선택 오류: {e}")

        await screenshot(page, "v2_teacher_click")

        # URL 변경 확인
        current_url = page.url
        print(f"  현재 URL: {current_url}")

        # 대시보드 로딩 대기
        try:
            await page.wait_for_selector("text=홈", timeout=10000)
            print("  대시보드 로드됨")
        except:
            print("  대시보드 로딩 실패, 현재 페이지 유지")

        await screenshot(page, "v2_teacher_dashboard")

        # 교사 테스트
        print("3. 교사 테스트...")
        teacher_tests = [
            ("홈", [("T01", "홈_기본", "홈"), ("T02", "홈_출석", "출석"), ("T03", "홈_과제", "과제")]),
            ("교과서", [("T04", "교과서_기본", "교과서")]),
            ("과제", [("T05", "과제_기본", "과제")]),
            ("AI 평가", [("T06", "AI평가_기본", "평가")]),
            ("AI 학습관", [("T07", "AI학습관_기본", "학습관")]),
            ("학생", [("T08", "학생_기본", "학생")]),
        ]

        for tab, tests in teacher_tests:
            await click_tab(page, tab)
            await page.wait_for_timeout(1000)
            for tid, tname, text in tests:
                r = "PASS" if await check(page, text, 2000) else "FAIL"
                results["teacher"][tid] = {"name": tname, "result": r}
                print(f"  {tname}: {r}")

        await screenshot(page, "v2_teacher_final")

        # 학생 전환
        print("4. 학생2 선택...")
        await page.goto(ENTRY_URL, timeout=60000)
        await patch_window_open(page)
        await page.wait_for_timeout(3000)

        try:
            selectors = [
                "text=학생 2",
                "text=학생2",
                "text=S2",
                "button:has-text('학생')",
            ]
            for sel in selectors:
                try:
                    if await page.is_visible(sel, timeout=2000):
                        await page.click(sel)
                        print(f"  클릭 성공: {sel}")
                        break
                except:
                    continue

            await page.wait_for_timeout(5000)
        except Exception as e:
            print(f"  학생 선택 오류: {e}")

        await screenshot(page, "v2_student_click")

        # 학생 테스트
        print("5. 학생 테스트...")
        student_tests = [
            ("홈", [("S01", "홈_기본", "홈"), ("S02", "홈_기분", "기분")]),
            ("교과서", [("S03", "교과서_기본", "교과서")]),
            ("과제", [("S04", "과제_기본", "과제")]),
            ("AI 평가", [("S05", "AI평가_기본", "평가")]),
            ("AI 학습관", [("S06", "AI학습관_기본", "학습관")]),
        ]

        for tab, tests in student_tests:
            await click_tab(page, tab)
            await page.wait_for_timeout(1000)
            for tid, tname, text in tests:
                r = "PASS" if await check(page, text, 2000) else "FAIL"
                results["student"][tid] = {"name": tname, "result": r}
                print(f"  {tname}: {r}")

        await screenshot(page, "v2_student_final")
        await browser.close()

def save_excel():
    today = datetime.now().strftime("%Y-%m-%d")
    wb = load_workbook(EXCEL_PATH)
    if today in wb.sheetnames:
        del wb[today]
    ws = wb.create_sheet(today)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = f"AIDT QA 테스트 ({today})"
    ws['A1'].font = Font(bold=True, size=14)

    row = 3
    headers = ['구분', '테스트명', '결과', '비고']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill, c.font, c.border = header_fill, header_font, border

    row = 4
    ws.cell(row=row, column=1, value="교사").font = Font(bold=True)
    for tid, data in results["teacher"].items():
        row += 1
        ws.cell(row=row, column=2, value=data['name']).border = border
        c = ws.cell(row=row, column=3, value=data['result'])
        c.border = border
        c.fill = pass_fill if data['result'] == 'PASS' else fail_fill

    row += 2
    ws.cell(row=row, column=1, value="학생").font = Font(bold=True)
    for tid, data in results["student"].items():
        row += 1
        ws.cell(row=row, column=2, value=data['name']).border = border
        c = ws.cell(row=row, column=3, value=data['result'])
        c.border = border
        c.fill = pass_fill if data['result'] == 'PASS' else fail_fill

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30

    wb.save(EXCEL_PATH)
    print(f"\n엑셀 저장 완료: {today}")
    return today

async def main():
    print("=" * 50)
    print("AIDT LMS QA 테스트 (window.open 패치)")
    print("=" * 50)

    await run_tests()

    # JSON 저장
    with open("D:/jaehyuk.myung/claude_demo/Demo_13_QA/quick_results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # 엑셀 저장
    try:
        save_excel()
    except Exception as e:
        print(f"엑셀 저장 실패: {e}")

    # 요약
    tp = sum(1 for r in results["teacher"].values() if r['result'] == 'PASS')
    tf = sum(1 for r in results["teacher"].values() if r['result'] == 'FAIL')
    sp = sum(1 for r in results["student"].values() if r['result'] == 'PASS')
    sf = sum(1 for r in results["student"].values() if r['result'] == 'FAIL')

    print("\n" + "=" * 50)
    print(f"교사: PASS {tp}, FAIL {tf}")
    print(f"학생: PASS {sp}, FAIL {sf}")
    print("=" * 50)

if __name__ == "__main__":
    asyncio.run(main())
